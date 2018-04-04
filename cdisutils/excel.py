import sys
import openpyxl
from io import BytesIO
from cdisutils.log import get_logger
from cdisutils.storage import BotoManager
from urlparse import urlparse

log = get_logger(
    "excel"
)

def combine_headers(headers):
    index = 0
    last_header = None
    new_header = []
    for value in headers[0]:
        if value:
            last_header = value
            if headers[1][index]:
                new_value = value + " " + headers[1][index]
            else:
                new_value = value
        else:
            if headers[1][index]:
                new_value = last_header + " " + headers[1][index]
            else:
                new_value = None
        if new_value:
            new_header.append(new_value)
        index += 1

    return new_header


def read_headers(rows, num_headers=1):
    """Read `num_headers` from rows and return combined headers

    Args:
        rows (generator): Rows generator from the worksheet
        num_headers (int): Number of headers to process

    Return:
        list: List of combined headers
    """
    header = None
    headers = []
    if num_headers > 2:
        log.info('Sorry, can only process 2 or fewer header rows')
        return header
    header_combine = num_headers > 1
    for row in rows:
        if num_headers > 0:
            header_data = []
            for cell in row:
                header_data.append(cell.value.strip()
                                   if cell.value.strip() else None)
            headers.append(header_data)
            num_headers -= 1
        if num_headers == 0:
            if header_combine:
                header = combine_headers(headers)
            else:
                header = headers[0]
            break
    return header


def read_columns(ws, num_headers=1):
    """Read an XLSX spreadsheet and return column values

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): Worksheet to read from
        num_headers (int): Number of headers in the work sheet

    Return:
        dict: Dictionary of lists, i.e. `{'key': [val1, val2]}`, `'key'` being
            column header name and the value being the list of column values
    """
    sheet_data = {}
    rows = ws.rows
    header = read_headers(rows, num_headers)
    if header is None:
        return sheet_data
    for row in rows:
        for h, cell in zip(header, row):
            value = (cell.value.strip()
                     if isinstance(cell.value, str) else cell.value)
            if h not in sheet_data:
                sheet_data[h] = [value]
            elif value:
                sheet_data[h].append(value)

    return sheet_data


def read_sheet(ws=None, num_headers=1):
    """Read an XLSX spreadsheet and return a list of rows

    Args:
        ws (openpyxl.worksheet.worksheet.Worksheet): Worksheet to read from
        num_headers (int): Number of headers in the worksheet. NOTE: maximum
            allowed number of headers is `2`

    Return:
        list: List of rows in format:

            sheet_data = [
                {
                    'header1': value11,
                    'header2': value12
                },
                {
                    'header1': value21,
                    'header2': value22
                }
            ]
    """
    sheet_data = []
    header = None
    if num_headers > 2:
        log.info("Sorry, can only process 2 or fewer header rows")
        return sheet_data

    if num_headers > 1:
        header_combine = True
    log.info("{} rows in sheet".format(ws.max_row))
    # NOTE: Saving the generator object in a variable, because accessing the
    # 'rows' property will always return a fresh generator
    rows = ws.rows

    header = read_headers(rows, num_headers)

    for row in rows:
        # NOTE: header rows will already be processed, remaining is just data
        # try and strip any blank rows
        any_data = False
        for cell in row:
            if cell.value and len(str(cell.value).strip()):
                any_data = True
                break

        if any_data:
            line_data = dict(zip(header, [str(cell.value).strip() for cell in row]))
            sheet_data.append(line_data)

    return sheet_data


def load_spreadsheet_from_s3(boto_man=None,
                             uri=None,
                             sheet_names=None,
                             num_headers=1):
    downloading = True
    total_transfer = 0
    data = None
    file_data = bytearray()
    chunk_size=16777216
    urlparts = urlparse(uri)
    s3_loc =  urlparts.netloc.split('.')[0]
    bucket = urlparts.path.split('/')[1]
    key_name = urlparts.path.split('/')[-1]



    # This is a bit tricky here. openpyxl likes a file, so we're
    # tricking it a bit by loading the data first, because we want to
    # stream from S3.
    data = boto_man.load_file(uri=uri)
    file_data.extend(data)

    try:
        wb = openpyxl.load_workbook(filename=BytesIO(file_data))
    except Exception as e:
        print 'Hmm, something wrong: {}'.format(e)
        if key_name.endswith('xls'):
            print 'Loading as xlsx'
            wb = openpyxl.load_workbook(filename=BytesIO(file_data),
                                        data_only=False)
        else:
            print 'Loading as xls'
            wb = openpyxl.load_workbook(filename=BytesIO(file_data),
                                        data_only=True)
    else:
        pass
    ws = None
    data = {}
    for name in wb.sheetnames:
        for sheet_name in sheet_names:
            if sheet_name in name:
                log.info("Found sheet: %s" % name)
                ws = wb[name]
                data[sheet_name] = read_sheet(ws=ws,
                                              num_headers=num_headers)

    return data


